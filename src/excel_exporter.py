from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def _header(ws, titles, row=1):
    for i, title in enumerate(titles, start=1):
        cell = ws.cell(row=row, column=i, value=title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(title) + 2)

def create_summary_sheet(wb, results):
    ws = wb.create_sheet("Summary", 0)
    _header(ws, ["Metric", "Value"])
    total_net = primer = finish = gallons = rolls = 0.0
    paint_rooms = wc_rooms = 0

    for r in results:
        total_net += r['net_area']
        m = r['materials']
        if m['type'] == 'paint':
            primer += m.get('primer_gallons', 0.0)
            finish += m.get('finish_gallons', 0.0)
            gallons += m.get('total_gallons', 0.0)
            paint_rooms += 1
        else:
            rolls += m.get('rolls', 0.0)
            wc_rooms += 1

    data = [
        ["Total Net Area (sf)", round(total_net, 2)],
        ["Paint – Primer Gallons", round(primer, 2)],
        ["Paint – Finish Gallons", round(finish, 2)],
        ["Paint – Total Gallons", round(gallons, 2)],
        ["Wallcovering – Rolls", round(rolls, 1)],
        ["# Paint Rooms", paint_rooms],
        ["# Wallcovering Rooms", wc_rooms],
    ]
    for i, (k, v) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

def create_room_breakdown_sheet(wb, results):
    ws = wb.create_sheet("Room Breakdown")
    _header(ws, [
        "Room ID","Name","L (ft)","W (ft)","H (ft)",
        "Doors","# Windows","Finish Type",
        "Gross Area (sf)","Net Area (sf)",
        "Primer (gal)","Finish (gal)","Total (gal)","Rolls"
    ])
    row = 2
    for r in results:
        room = r['room']
        m = r['materials']
        ws.cell(row=row, column=1, value=room['id'])
        ws.cell(row=row, column=2, value=room['name'])
        ws.cell(row=row, column=3, value=room['length'])
        ws.cell(row=row, column=4, value=room['width'])
        ws.cell(row=row, column=5, value=room['height'])
        ws.cell(row=row, column=6, value=room['doors'])
        ws.cell(row=row, column=7, value=room['windows'])
        ws.cell(row=row, column=8, value=m['type'])
        ws.cell(row=row, column=9, value=r['gross_area'])
        ws.cell(row=row, column=10, value=r['net_area'])
        ws.cell(row=row, column=11, value=m.get('primer_gallons'))
        ws.cell(row=row, column=12, value=m.get('finish_gallons'))
        ws.cell(row=row, column=13, value=m.get('total_gallons'))
        ws.cell(row=row, column=14, value=m.get('rolls'))
        row += 1

def create_system_breakdown_sheet(wb, results):
    ws = wb.create_sheet("System Breakdown")
    _header(ws, ["System","Rooms","#","Net Area (sf)","Primer (gal)","Finish (gal)","Total (gal)","Rolls"])

    agg = {
        "paint": {"rooms": 0, "area": 0.0, "primer": 0.0, "finish": 0.0, "total": 0.0},
        "wallcovering": {"rooms": 0, "area": 0.0, "rolls": 0.0},
    }
    for r in results:
        m = r['materials']
        if m['type'] == 'paint':
            agg["paint"]["rooms"] += 1
            agg["paint"]["area"] += r['net_area']
            agg["paint"]["primer"] += m.get('primer_gallons', 0.0)
            agg["paint"]["finish"] += m.get('finish_gallons', 0.0)
            agg["paint"]["total"] += m.get('total_gallons', 0.0)
        else:
            agg["wallcovering"]["rooms"] += 1
            agg["wallcovering"]["area"] += r['net_area']
            agg["wallcovering"]["rolls"] += m.get('rolls', 0.0)

    ws.append(["paint", "rooms", agg["paint"]["rooms"], round(agg["paint"]["area"],2),
               round(agg["paint"]["primer"],2), round(agg["paint"]["finish"],2),
               round(agg["paint"]["total"],2), None])
    ws.append(["wallcovering", "rooms", agg["wallcovering"]["rooms"], round(agg["wallcovering"]["area"],2),
               None, None, None, round(agg["wallcovering"]["rolls"],1)])

def generate_workbook(results, output_path):
    wb = Workbook()
    # remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    create_summary_sheet(wb, results)
    create_room_breakdown_sheet(wb, results)
    create_system_breakdown_sheet(wb, results)
    wb.save(output_path)
    return output_path
